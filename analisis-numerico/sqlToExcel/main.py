import sqlparse
import re
from openpyxl import Workbook
import logging

logging.basicConfig(level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')

def parse_create_table(statement):
    logging.debug(f"Parsing CREATE statement: {statement[:100]}...")
    table_match = re.search(r'CREATE TABLE `?(\w+)`?', statement, re.IGNORECASE)
    if not table_match:
        logging.warning(f"No se pudo encontrar el nombre de la tabla en: {statement[:100]}...")
        return None, None
    table_name = table_match.group(1)

    columns = []
    lines = statement.split('\n')
    for line in lines:
        line = line.strip()
        if line.startswith('`'):
            parts = line.split('`')
            column_name = parts[1]
            columns.append(column_name)

    logging.debug(f"Tabla encontrada: {table_name}, Columnas: {columns}")
    return table_name, columns

def parse_value(value):
    value = value.strip()
    if value.upper() == 'NULL':
        return None
    elif value.startswith("'") and value.endswith("'"):
        return value[1:-1].replace("''", "'")
    elif value.isdigit():
        return int(value)
    elif re.match(r'^-?\d+(\.\d+)?$', value):
        return float(value)
    else:
        return value

def parse_insert(statement):
    logging.debug(f"Parsing INSERT statement: {statement[:100]}...")
    table_match = re.search(r'INSERT INTO `?(\w+)`?\s*(?:$$[^)]+$$)?\s*VALUES?\s*', statement, re.IGNORECASE)
    if not table_match:
        logging.warning(f"No se pudo encontrar el nombre de la tabla en INSERT: {statement[:100]}...")
        return None, []
    table_name = table_match.group(1)

    values_pattern = r'$$([^()]+(?:\([^()]*$$[^()]*)*)\)'
    values_match = re.findall(values_pattern, statement)
    values = []
    for value_string in values_match:
        row = []
        current_value = ''
        in_quotes = False
        for char in value_string:
            if char == "'" and not in_quotes:
                in_quotes = True
                current_value += char
            elif char == "'" and in_quotes:
                in_quotes = False
                current_value += char
            elif char == ',' and not in_quotes:
                row.append(parse_value(current_value))
                current_value = ''
            else:
                current_value += char
        if current_value:
            row.append(parse_value(current_value))
        values.append(row)

    logging.debug(f"Insertando en tabla: {table_name}, Número de filas: {len(values)}")
    return table_name, values

def sql_to_excel(sql_file, excel_file):
    logging.info(f"Iniciando conversión de {sql_file} a {excel_file}")
    try:
        with open(sql_file, 'r', encoding='latin-1') as file:
            sql = file.read()
        logging.info(f"Archivo SQL leído correctamente. Tamaño: {len(sql)} caracteres")
    except Exception as e:
        logging.error(f"Error al leer el archivo SQL: {e}")
        return

    statements = sqlparse.split(sql)
    logging.info(f"Número de declaraciones SQL encontradas: {len(statements)}")

    wb = Workbook()
    tables_processed = 0
    table_columns = {}

    for statement in statements:
        parsed = sqlparse.parse(statement)[0]

        if parsed.get_type() == 'CREATE':
            table_name, columns = parse_create_table(statement)
            if table_name and columns:
                table_columns[table_name] = columns
                ws = wb.create_sheet(title=table_name[:31])  # Excel limita nombres de hojas a 31 caracteres
                ws.append(columns)
                tables_processed += 1
                logging.info(f"Tabla creada en Excel: {table_name}")

        elif parsed.get_type() == 'INSERT':
            table_name, values = parse_insert(statement)
            if table_name in table_columns:
                ws = wb[table_name[:31]]
                columns = table_columns[table_name]

                for row in values:
                    row_data = row + [None] * (len(columns) - len(row))
                    ws.append(row_data)

                logging.info(f"Datos insertados en tabla: {table_name}, Filas: {len(values)}")

    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    logging.info(f"Tablas procesadas: {tables_processed}")

    try:
        wb.save(excel_file)
        logging.info(f"Archivo Excel guardado exitosamente: {excel_file}")
    except Exception as e:
        logging.error(f"Error al guardar el archivo Excel: {e}")

# Uso del script
sql_to_excel('/Users/tomasmatteozzi/Desktop/extranet.sql', 'resultado_completo.xlsx')
logging.info("Script completado")